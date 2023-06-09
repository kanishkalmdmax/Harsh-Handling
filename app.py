from flask import Flask, request, send_file
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
import os

app = Flask(__name__)

@app.route('/')
def index():
    return '''
    <!doctype html>
    <html>
        <head>
            <title>Netradyne Harsh Handling E-Mail</title>
            <style>
                    #instructions {
                        border: 1px solid black;
                        padding: 10px;
                        position: absolute;
                        top: 50%;
                        right: 10px;
                        transform: translateY(-50%);
                    }
            </style>
        </head>
        <body>
            <h3>Netradyne Harsh Handling E-Mail Grid (Daily)</h3>
            <form action="/upload" method="post" enctype="multipart/form-data">
                <input type="file" name="file" accept=".xlsx"><br><br>
                <input type="submit" value="Upload">
            </form>
                <div id="instructions">
                    Steps:<br>(Upload Reports in the Performance App first)<br>1. Open Performance App, then select Reports and then Driver Report.<br>2. Click on the Daily button and select yesterday's date (or as per preference), click Download.<br>3. Once the file is downloaded, open this URL, click on Choose File, select the file, and click Upload.<br>4. Then a new page will open with the Download button, click on that and download the file.<br>5. This file will have the email grid, you can copy this grid and paste it in the email.<br>6. If you see an Error Message, it would mean you either uploaded an incorrect file,<br>perhaps you accidently clicked the Summary option, or directly tried to upload the CSV file downloaded from Netradyne<br>or uploaded something else entirely.
                </div>
        </body>
    </html>
    '''

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['file']
    filename = secure_filename(file.filename)

    # Check if the uploaded file is an excel file
    if not filename.endswith('.xlsx'):
        return '''
        <!doctype html>
        <html>
            <head>
                <title>Netradyne Harsh Handling E-Mail</title>
            </head>
            <body>
                <h3>Error: This is not a XLSX file</h3>
                <p>This is not a XLSX file, please download &amp; upload the file from Performance App&gt;Driver Report&gt;Summary</p>
                <a href="/">Return to upload form</a>
            </body]
        </html>
        '''

    file.save(filename)

    # Load the excel file
    try:
        df = pd.read_excel(filename)
    except Exception as e:
        return '''
        <!doctype html>
        <html>
            <head>
                <title>Netradyne Harsh Handling E-Mail</title>
            </head>
            <body>
                <h3>Error: This is not a valid Driver Report file</h3>
                <p>This is not a valid Driver Report file, please download the correct file, please download &amp; upload the file from Performance App&gt;Driver Report&gt;Summary</p>
                <a href="/">Return to upload form</a>
            </body>
        </html>
        '''

    # Define the columns to extract data from
    columns = ['Name', 'Following Distance', 'Camera Obstruction', 'U Turn', 'Driver Distraction', 'Seatbelt Compliance', 'Sign Violations', 'Speeding Violations', 'Traffic Light Violation', 'Hard Turn', 'Hard Braking', 'Hard Acceleration']

    # Check if the excel file has the required columns
    if not all(column in df.columns for column in columns):
        return '''
        <!doctype html>
        <html>
            <head>
                <title>Netradyne Harsh Handling E-Mail</title>
            </head>
            <body>
                <h3>Error: This is not a valid Driver Report file</h3>
                <p>This is not a valid Driver Report file, please download the correct file, please download &amp; upload the file from Performance App&gt;Driver Report&gt;Summary</p>
                <a href="/">Return to upload form</a>
            </body>
        </html>
        '''

    # Filter the rows based on the condition
    df = df[(df['Following Distance'] > 0) | (df['Camera Obstruction'] > 0) | (df['U Turn'] > 0) | (df['Driver Distraction'] > 0) | (df['Seatbelt Compliance'] > 0) | (df['Sign Violations'] > 0) | (df['Speeding Violations'] > 0) | (df['Traffic Light Violation'] > 0) | (df['Hard Acceleration'] > 0) | (df['Hard Turn'] > 0) | (df['Hard Braking'] > 0)]

    # Group by name and aggregate violations
    df = df.groupby('Name').agg({
        'Following Distance': lambda x: sum(val for val in x if val >= 1),
        'Camera Obstruction': lambda x: sum(val for val in x if val >= 1),
        'U Turn': lambda x: sum(val for val in x if val >= 1),
        'Driver Distraction': lambda x: sum(val for val in x if val >= 1),
        'Seatbelt Compliance': lambda x: sum(val for val in x if val >= 1),
        'Sign Violations': lambda x: sum(val for val in x if val >= 1),
        'Speeding Violations': lambda x: sum(val for val in x if val >= 1),
        'Traffic Light Violation': lambda x: sum(val for val in x if val >= 1),
        'Hard Turn': lambda x: sum(val for val in x if val >= 1),
        'Hard Braking': lambda x: sum(val for val in x if val >= 1),
        'Hard Acceleration': lambda x: sum(val for val in x if val >= 1)
    }).reset_index()

    # Create a new column for Violations
    df['Violations'] = df.apply(lambda row: ', '.join([f"{col} ({val})" for col, val in row.items() if col in ['Hard Turn', 'Hard Braking', 'Hard Acceleration', 'U Turn', 'Sign Violations', 'Speeding Violations', 'Traffic Light Violation', 'Following Distance', 'Camera Obstruction', 'Driver Distraction', 'Seatbelt Compliance'] and val >= 1]), axis=1)


    # Create a new column for Violations Count
    df['Violations Count'] = df[columns[1:]].apply(lambda x: x[x>0].sum(), axis=1)

    # Select the required columns
    df = df[['Name', 'Violations', 'Violations Count']]

    # Write the output to a new excel file
    output_file = filename.replace('.xlsx', '_grid.xlsx')
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, sheet_name='Extracted Data', index=False)

    # Apply formatting to the cells
    workbook = writer.book
    worksheet = writer.sheets['Extracted Data']

    # Set alignment to center and middle for all cells
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set fill color for cells A1:C1
    fill = PatternFill(fill_type='solid', fgColor='B8CCE4')
    for cell in worksheet['A1:C1'][0]:
        cell.fill = fill
        
    # Apply border to all cells
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border

    # Adjust column width to fit data
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.1
        worksheet.column_dimensions[column_letter].width = adjusted_width

    writer.close()

    # Return response with download button
    return '''
    <!doctype html>
    <html>
        <head>
            <title>Netradyne Harsh Handling E-Mail</title>
        </head>
        <body>
            <h3>Netradyne Harsh Handling E-Mail Grid (Daily)</h3>
            <p>File processed successfully!</p>
            <a href="/download/{}"><button>Download</button></a>
        </body>
    </html>
    '''.format(output_file)

@app.route('/download/<path:filename>')
def download(filename):
    return send_file(filename, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=os.environ.get('PORT', 5000),debug=True)
